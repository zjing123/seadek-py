#!/usr/bin/env python3.6
# -*- coding:UTF-8 -*-
# @TIME   : 19-3-13 下午3:40
# @Author : Liuchuan
# @File   : parse.py

from os import path
import openpyxl

MAX_ROW = 272
MAX_COL = 10

filename = path.dirname(path.abspath(__file__)) + '/product-info.xlsx'

inwb = openpyxl.load_workbook(filename)
sheet_names = inwb.get_sheet_names()

ws = inwb.get_sheet_by_name(sheet_names[0])
rows = ws.max_row
cols = ws.max_column

products = []
for r in range(1, MAX_ROW + 1):
    product = []
    for c in range(1, MAX_COL + 1):
        product.append(ws.cell(r,c).value)
    if r == 1:
        continue
    products.append(product)

print(products[0])

# print(ws.max_row, ws.max_column)
